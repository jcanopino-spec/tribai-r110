"""Inspeccion exhaustiva de las hojas de configuracion del .xlsm.

Recolecta todas las celdas con texto/formulas/listas desplegables que el
.xlsm usa como parametros del calculo.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook


SHEETS_TO_INSPECT = [
    "Datos Básicos",
    "Datos Informativos",
    "Inicio",
    "Sanciones",
    "Anexo 23 Beneficio Aud.",
    "Anexo 12 Deterioro Cartera",
    "Anexo 7 Venta de Acciones",
    "Anexo 8 Ganancia Ocasional",
    "Anexo 13 Ded IVA Bienes Cap.",
    "Anexo 19 Rentas Exentas",
    "Anexo 25 Cálculo Dividendos ",
    "Inversiones en Asoc",
]


def main():
    src = Path(sys.argv[1])
    wb = load_workbook(src, data_only=False, keep_links=False)

    for name in SHEETS_TO_INSPECT:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        print(f"\n{'='*60}")
        print(f"=== {name}  ({ws.max_row}r x {ws.max_column}c)  state={ws.sheet_state}")
        print('='*60)

        # Show non-empty cells with text labels
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 100), max_col=min(ws.max_column, 12), values_only=False):
            row_text = []
            has_content = False
            for cell in row:
                if cell.value is not None:
                    has_content = True
                    v = cell.value
                    if isinstance(v, str):
                        v_str = v.strip()
                        if v_str.startswith('='):
                            row_text.append(f"{cell.coordinate}=[F]")
                        else:
                            row_text.append(f"{cell.coordinate}={v_str[:50]}")
                    else:
                        row_text.append(f"{cell.coordinate}={v}")
                else:
                    row_text.append('')
            if has_content:
                print('  ' + ' | '.join([t for t in row_text if t]))

        # Show data validations (dropdowns)
        if ws.data_validations.dataValidation:
            print(f'  --- Validaciones (dropdowns) ---')
            for dv in ws.data_validations.dataValidation:
                f1 = (dv.formula1 or '')[:80]
                ranges = ', '.join([str(r) for r in dv.sqref.ranges]) if dv.sqref else ''
                print(f'    type={dv.type} f1={f1} ranges={ranges}')


if __name__ == "__main__":
    main()
