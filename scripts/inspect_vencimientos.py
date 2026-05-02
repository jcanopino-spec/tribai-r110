"""Inspecciona la hoja Vencimientos completa."""

from __future__ import annotations
import sys
from pathlib import Path
from openpyxl import load_workbook


def main():
    src = Path(sys.argv[1])
    wb = load_workbook(src, data_only=True, keep_links=False)
    ws = wb["Vencimientos"]
    print(f"{ws.max_row}r x {ws.max_column}c\n")
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
        cells = []
        for cell in r:
            v = cell.value
            if v is not None:
                if hasattr(v, "isoformat"):
                    v = v.strftime("%Y-%m-%d")
                cells.append(f"{cell.coordinate}={str(v)[:40]}")
        if cells:
            print("  " + " | ".join(cells))


if __name__ == "__main__":
    main()
