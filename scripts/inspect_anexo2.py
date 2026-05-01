"""Inspecciona Anexo 2 (Anticipo) del .xlsm fuente."""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook


def main():
    src = Path(sys.argv[1])
    wb = load_workbook(src, data_only=False, keep_links=False)
    ws = wb["Anexo 2 Anticipo"]
    print(f"Sheet: {ws.title}  ({ws.max_row}r x {ws.max_column}c)\n")
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=8, values_only=False):
        for cell in r:
            v = cell.value
            if v is not None:
                v_str = str(v)
                if v_str.startswith("="):
                    print(f"  {cell.coordinate:>5}: {v_str[:150]}")
                else:
                    print(f"  {cell.coordinate:>5}: \"{v_str[:80]}\"")


if __name__ == "__main__":
    main()
