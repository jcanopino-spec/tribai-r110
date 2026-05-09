#!/usr/bin/env python3
"""Carga las 6 declaraciones de IVA bimestrales de SISTEMAS ARIES SAS
del año gravable 2025 al anexo de IVA de la app.

Lee:
  - Excel resumen con los 6 bimestres
  - PDFs oficiales del F300 (B1-B6)

Inserta:
  - 6 registros en anexo_iva_declaraciones
  - 6 PDFs en bucket Supabase Storage 'anexo-iva-pdfs'

Pre-requisitos:
  · Migración 028 aplicada (anexo_iva_declaraciones)
  · Bucket 'anexo-iva-pdfs' creado (lo crea automáticamente si no existe)

Uso:
    python3 scripts/seed_iva_aries.py
"""

import json
import ssl
import urllib.parse
import urllib.request
from pathlib import Path

import certifi
from openpyxl import load_workbook

# ============================================================
# CONFIGURACIÓN
# ============================================================
ROOT = Path(__file__).resolve().parent.parent
SOURCE = Path(
    "/Users/jacp/Library/CloudStorage/OneDrive-Personal/CLAUDE/RENTA_110/110X/PREVIO ARIES/ivas2025sistemasaries"
)
EXCEL_PATH = SOURCE / "IVA ANEXO ARIES VF 2025v-2.xlsx"

# IDs ya conocidos · obtenidos via REST query
EMPRESA_NIT = "800210193"  # SISTEMAS ARIES SAS
DECLARACION_ID = "6a66c036-904d-4953-8559-38a42ff4909e"  # AG 2025

# Bucket de Storage para PDFs
BUCKET = "anexo-iva-pdfs"

# Casillas del F300 que mapean a nuestro schema
CASILLAS_F300 = {
    "ingresos_brutos": 39,
    "devoluciones": 40,
    # ingresos_netos no se guarda · se calcula
    "ingresos_no_gravados": 38,  # casilla 38 "Por operaciones no gravadas"
    "ingresos_exentos": 35,  # "Por operaciones exentas (art. 477, 478, 481)"
    "ingresos_gravados_general": 28,  # "Tarifa general"
    "ingresos_gravados_5": 27,  # "Tarifa 5%"
    "iva_generado": 63,  # "Total impuesto generado"
    "iva_descontable": 77,  # "Total impuestos descontables"
    "saldo_pagar": 78,  # "Saldo a pagar por el periodo fiscal"
    "saldo_favor": 79,  # "Saldo a favor del periodo fiscal"
}

# Cada bimestre está en una columna del Excel
BIM_COLS = {1: 9, 2: 10, 3: 11, 4: 12, 5: 13, 6: 14}

# PDFs por bimestre
PDF_FILES = {
    1: SOURCE / "IVA B1 2025.pdf",
    2: SOURCE / "IVA B2 2025.pdf",
    3: SOURCE / "IVA B3 2025.pdf",
    4: SOURCE / "IVA B4 2025.pdf",
    5: SOURCE / "IVA B5 2025.pdf",
    6: SOURCE / "IVA B6 2025.pdf",
}


# ============================================================
# UTILIDADES
# ============================================================
def load_env() -> dict:
    env = {}
    for line in (ROOT / ".env.local").read_text().splitlines():
        if "=" in line and not line.startswith("#"):
            k, _, v = line.partition("=")
            env[k.strip()] = v.strip().strip("'\"")
    return env


def num(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def http_request(
    url: str, method: str, headers: dict, data: bytes | None = None
) -> tuple[int, bytes]:
    ctx = ssl.create_default_context(cafile=certifi.where())
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        r = urllib.request.urlopen(req, context=ctx, timeout=30)
        return r.status, r.read()
    except urllib.error.HTTPError as e:
        return e.code, e.read()


# ============================================================
# EXTRACCIÓN DEL EXCEL
# ============================================================
def extraer_bimestres() -> list[dict]:
    print(f"📖 Leyendo {EXCEL_PATH.name}...")
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["IVA 2025"]

    # Mapear casilla → fila
    casilla_a_fila = {}
    for r in range(1, ws.max_row + 1):
        cas = ws.cell(r, 8).value
        try:
            casilla_a_fila[int(cas)] = r
        except (ValueError, TypeError):
            pass

    bimestres = []
    for periodo, col in BIM_COLS.items():
        def get(casilla: int) -> float:
            fila = casilla_a_fila.get(casilla)
            if not fila:
                return 0.0
            return num(ws.cell(fila, col).value)

        ingresos_brutos = get(CASILLAS_F300["ingresos_brutos"])
        gravados = get(CASILLAS_F300["ingresos_gravados_general"]) + get(
            CASILLAS_F300["ingresos_gravados_5"]
        )

        bim = {
            "periodicidad": "bimestral",
            "periodo": periodo,
            "fecha_presentacion": None,  # No está en el Excel
            "numero_formulario": None,
            "ingresos_brutos": ingresos_brutos,
            "ingresos_no_gravados": get(CASILLAS_F300["ingresos_no_gravados"]),
            "ingresos_exentos": get(CASILLAS_F300["ingresos_exentos"]),
            "ingresos_gravados": gravados,
            "iva_generado": get(CASILLAS_F300["iva_generado"]),
            "iva_descontable": get(CASILLAS_F300["iva_descontable"]),
            "saldo_pagar": get(CASILLAS_F300["saldo_pagar"]),
            "saldo_favor": get(CASILLAS_F300["saldo_favor"]),
            "observacion": "Cargado desde IVA ANEXO ARIES VF 2025v-2.xlsx",
        }
        bimestres.append(bim)
        print(
            f"  Bim {periodo}: ingresos {ingresos_brutos:>16,.0f} · saldo pagar {bim['saldo_pagar']:>14,.0f}"
        )
    return bimestres


# ============================================================
# UPLOAD PDF
# ============================================================
def upload_pdf(env: dict, periodo: int, pdf_path: Path) -> str | None:
    if not pdf_path.exists():
        print(f"  ⚠ PDF no encontrado: {pdf_path}")
        return None

    storage_path = f"{DECLARACION_ID}/bimestral-{periodo}-{pdf_path.name.replace(' ', '_')}"
    pdf_bytes = pdf_path.read_bytes()

    upload_url = (
        f"{env['NEXT_PUBLIC_SUPABASE_URL']}/storage/v1/object/{BUCKET}/{storage_path}"
    )
    headers = {
        "apikey": env["SUPABASE_SERVICE_ROLE_KEY"],
        "Authorization": f"Bearer {env['SUPABASE_SERVICE_ROLE_KEY']}",
        "Content-Type": "application/pdf",
        "x-upsert": "true",
    }
    status, body = http_request(upload_url, "POST", headers, pdf_bytes)
    if status >= 400:
        print(f"  ✗ Upload PDF falló ({status}): {body.decode()[:120]}")
        return None
    return storage_path


# ============================================================
# UPSERT DE REGISTROS
# ============================================================
def upsert_iva(env: dict, bim: dict) -> bool:
    payload = {**bim, "declaracion_id": DECLARACION_ID}
    payload_bytes = json.dumps(payload).encode()
    upsert_url = (
        f"{env['NEXT_PUBLIC_SUPABASE_URL']}/rest/v1/anexo_iva_declaraciones"
        "?on_conflict=declaracion_id,periodicidad,periodo"
    )
    headers = {
        "apikey": env["SUPABASE_SERVICE_ROLE_KEY"],
        "Authorization": f"Bearer {env['SUPABASE_SERVICE_ROLE_KEY']}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates,return=representation",
    }
    status, body = http_request(upsert_url, "POST", headers, payload_bytes)
    if status >= 400:
        print(f"  ✗ Upsert falló ({status}): {body.decode()[:200]}")
        return False
    return True


def update_pdf_path(
    env: dict, periodo: int, pdf_path: str, pdf_filename: str
) -> bool:
    qs = urllib.parse.urlencode(
        {
            "declaracion_id": f"eq.{DECLARACION_ID}",
            "periodicidad": "eq.bimestral",
            "periodo": f"eq.{periodo}",
        }
    )
    update_url = (
        f"{env['NEXT_PUBLIC_SUPABASE_URL']}/rest/v1/anexo_iva_declaraciones?{qs}"
    )
    payload = json.dumps({"pdf_path": pdf_path, "pdf_filename": pdf_filename}).encode()
    headers = {
        "apikey": env["SUPABASE_SERVICE_ROLE_KEY"],
        "Authorization": f"Bearer {env['SUPABASE_SERVICE_ROLE_KEY']}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }
    status, body = http_request(update_url, "PATCH", headers, payload)
    if status >= 400:
        print(f"  ✗ Update PDF path falló ({status}): {body.decode()[:120]}")
        return False
    return True


# ============================================================
# MAIN
# ============================================================
def main():
    env = load_env()

    # Verificar que la tabla existe
    check_url = (
        f"{env['NEXT_PUBLIC_SUPABASE_URL']}/rest/v1/anexo_iva_declaraciones?limit=1"
    )
    headers = {
        "apikey": env["SUPABASE_SERVICE_ROLE_KEY"],
        "Authorization": f"Bearer {env['SUPABASE_SERVICE_ROLE_KEY']}",
    }
    status, _ = http_request(check_url, "GET", headers)
    if status >= 400:
        print(
            "\n✗ Tabla anexo_iva_declaraciones no existe. Aplica primero la "
            "migración 028:"
        )
        print(
            "  supabase/migrations/20260509_028_anexo_iva.sql en SQL Editor "
            "del dashboard Supabase\n"
        )
        return

    print(f"\n🎯 Empresa: SISTEMAS ARIES SAS · NIT {EMPRESA_NIT}")
    print(f"   Declaración: AG 2025 · {DECLARACION_ID}")
    print(f"   Bucket: {BUCKET}\n")

    # 1. Extraer datos del Excel
    bimestres = extraer_bimestres()
    total_ingresos = sum(b["ingresos_brutos"] for b in bimestres)
    total_pagar = sum(b["saldo_pagar"] for b in bimestres)
    print(f"\n📊 TOTAL año:")
    print(f"   Ingresos brutos: ${total_ingresos:,.0f}")
    print(f"   Saldo a pagar:   ${total_pagar:,.0f}")

    # 2. Insertar registros
    print("\n💾 Insertando registros en BD...")
    for bim in bimestres:
        ok = upsert_iva(env, bim)
        if ok:
            print(f"  ✓ Bim {bim['periodo']} guardado")

    # 3. Subir PDFs
    print("\n📎 Subiendo PDFs al bucket...")
    for periodo, pdf_path in PDF_FILES.items():
        if not pdf_path.exists():
            continue
        storage_path = upload_pdf(env, periodo, pdf_path)
        if storage_path:
            update_pdf_path(env, periodo, storage_path, pdf_path.name)
            print(f"  ✓ Bim {periodo}: {pdf_path.name}")

    print("\n✓ Carga completa.")


if __name__ == "__main__":
    main()
