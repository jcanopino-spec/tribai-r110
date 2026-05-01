"""Inspeccion de las hojas del formulario 110 para identificar la estructura
de renglones (codigo, descripcion, formula).

Vuelca un dump de las primeras 25 columnas de cada hoja a JSON para que podamos
analizarlo manualmente y luego construir un extractor preciso.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


SHEETS = ["110 MUISCA", "Formulario 110"]
MAX_COL = 25


def dump_sheet(ws, max_col: int = MAX_COL) -> dict:
    rows = []
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=max_col, values_only=False):
        row_data = []
        for cell in r:
            if cell.value is None:
                row_data.append(None)
            else:
                v = cell.value
                if isinstance(v, str):
                    v = v.strip()
                row_data.append(v)
        # skip rows where all are None
        if all(c is None or c == "" for c in row_data):
            row_data = None
        rows.append(row_data)
    return {
        "title": ws.title,
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "rows": rows,
    }


def main():
    if len(sys.argv) != 2:
        print("usage: inspect_form110.py <path>", file=sys.stderr)
        sys.exit(1)
    src = Path(sys.argv[1])
    out_dir = Path(__file__).resolve().parent.parent / "data" / "extracted"
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(src, data_only=False, keep_links=False)
    out = {}
    for name in SHEETS:
        if name in wb.sheetnames:
            out[name] = dump_sheet(wb[name])
            print(f"{name}: {out[name]['max_row']} rows x {out[name]['max_col']} cols (dump capped at {MAX_COL})")
        else:
            print(f"WARN: sheet '{name}' not found")

    out_path = out_dir / "_form110_inspection.json"
    out_path.write_text(json.dumps(out, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
    print(f"Dump → {out_path}")


if __name__ == "__main__":
    main()
