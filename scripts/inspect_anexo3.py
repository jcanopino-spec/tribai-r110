"""Inspecciona Anexo 3 Retenciones y Autorretenciones."""

from __future__ import annotations
import sys
from pathlib import Path
from openpyxl import load_workbook


def main():
    src = Path(sys.argv[1])
    wb = load_workbook(src, data_only=False, keep_links=False)
    ws = wb["Anexo 3 Retenciones y Auto"]
    print(f"{ws.max_row}r x {ws.max_column}c\n")

    for r in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 140), max_col=min(ws.max_column, 16), values_only=False):
        cells = []
        for cell in r:
            v = cell.value
            if v is not None:
                if hasattr(v, "isoformat"):
                    v = v.strftime("%Y-%m-%d")
                s = str(v).strip()
                if s.startswith("="):
                    cells.append(f"{cell.coordinate}=[F]")
                else:
                    cells.append(f"{cell.coordinate}={s[:50]}")
        if cells:
            print("  " + " | ".join(cells))


if __name__ == "__main__":
    main()
