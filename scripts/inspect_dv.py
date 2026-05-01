"""Buscar como el .xlsm calcula el DV del NIT.

Lo busca en la hoja 'Datos Informativos' (donde D13 es el DV) y muestra la
formula y celdas referenciadas para auditarla contra el algoritmo estandar DIAN.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook


def main():
    src = Path(sys.argv[1])
    wb = load_workbook(src, data_only=False, keep_links=False)
    ws = wb["Datos Informativos"]

    # Print formulas around D12-D13 (NIT and DV) and any defined names with 'dv' or 'nit'
    print("=== Datos Informativos rows 10-20 (formulas) ===")
    for r in ws.iter_rows(min_row=10, max_row=20, max_col=8, values_only=False):
        for cell in r:
            if cell.value is not None:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    print(f"  {cell.coordinate}: {v}")
                elif isinstance(v, str):
                    print(f"  {cell.coordinate}: \"{v[:80]}\"")
                else:
                    print(f"  {cell.coordinate}: {v!r}")

    print()
    print("=== Defined names containing 'dv', 'nit', 'verif' ===")
    for n in wb.defined_names:
        nm = n.lower()
        if any(t in nm for t in ("dv", "nit", "verif", "digito")):
            v = wb.defined_names[n].value
            print(f"  {n}: {v}")


if __name__ == "__main__":
    main()
