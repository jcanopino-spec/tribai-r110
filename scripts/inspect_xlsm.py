"""Inspeccion de alto nivel del .xlsm fuente.

Lista hojas, dimensiones, nombres definidos (named ranges), validaciones de datos
(listas desplegables) y un sample de las primeras filas de cada hoja.

Uso:
    python3 scripts/inspect_xlsm.py "<ruta al .xlsm>"

Salida:
    data/extracted/00_inspection.json
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def inspect(xlsm_path: Path) -> dict:
    wb = load_workbook(xlsm_path, data_only=False, keep_links=False)

    sheets = []
    for ws in wb.worksheets:
        dvs = []
        for dv in ws.data_validations.dataValidation:
            dvs.append({
                "type": dv.type,
                "formula1": dv.formula1,
                "formula2": dv.formula2,
                "ranges": [str(r) for r in dv.sqref.ranges] if dv.sqref else [],
                "allow_blank": dv.allowBlank,
                "show_dropdown": dv.showDropDown,
            })

        sheets.append({
            "title": ws.title,
            "state": ws.sheet_state,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "dimension": ws.dimensions,
            "data_validations": dvs,
            "merged_cells_count": len(ws.merged_cells.ranges),
        })

    defined_names = []
    for dn in wb.defined_names:
        try:
            destinations = list(wb.defined_names[dn].destinations)
        except Exception:
            destinations = []
        defined_names.append({
            "name": dn,
            "value": wb.defined_names[dn].value,
            "destinations": [
                {"sheet": s, "range": r} for (s, r) in destinations
            ],
        })

    return {
        "file": str(xlsm_path),
        "sheets_count": len(sheets),
        "sheets": sheets,
        "defined_names_count": len(defined_names),
        "defined_names": defined_names,
        "props": {
            "title": wb.properties.title,
            "creator": wb.properties.creator,
            "modified": str(wb.properties.modified) if wb.properties.modified else None,
            "company": getattr(wb.properties, "company", None),
        },
    }


def main():
    if len(sys.argv) != 2:
        print("usage: inspect_xlsm.py <path>", file=sys.stderr)
        sys.exit(1)
    src = Path(sys.argv[1])
    out_dir = Path(__file__).resolve().parent.parent / "data" / "extracted"
    out_dir.mkdir(parents=True, exist_ok=True)
    report = inspect(src)
    out_path = out_dir / "00_inspection.json"
    out_path.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")

    print(f"Hojas: {report['sheets_count']}")
    print(f"Nombres definidos: {report['defined_names_count']}")
    print(f"Reporte → {out_path}")
    print()
    print("Hojas:")
    for s in report["sheets"]:
        flag = "" if s["state"] == "visible" else f"  [{s['state']}]"
        dvs = len(s["data_validations"])
        print(f"  - {s['title']:<45} {s['max_row']:>5}r x {s['max_col']:>3}c   dv={dvs}{flag}")


if __name__ == "__main__":
    main()
