"""Inspeccion del Balance de Prueba: columnas esperadas y muestra de filas.

La hoja tiene 2507 filas. Vemos el header y un sample para entender estructura.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook


def main():
    src = Path(sys.argv[1])
    wb = load_workbook(src, data_only=True, keep_links=False)

    for name in ["Balance de Prueba", "Hoja Sumaria", "Sanciones", "Vencimientos"]:
        if name not in wb.sheetnames:
            print(f"WARN: {name} not found")
            continue
        ws = wb[name]
        print(f"\n=== {name}  ({ws.max_row}r x {ws.max_column}c) ===")
        # Show first 8 non-empty rows
        shown = 0
        for r in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in r):
                continue
            cells = [str(v)[:35] if v is not None else "" for v in r]
            print("  " + " | ".join(cells[:12]))
            shown += 1
            if shown >= 8:
                break

        # Then show last 3 non-empty rows
        last_rows = []
        for r in ws.iter_rows(min_row=max(1, ws.max_row - 30), max_row=ws.max_row, values_only=True):
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in r):
                continue
            cells = [str(v)[:35] if v is not None else "" for v in r]
            last_rows.append("  " + " | ".join(cells[:12]))
        print("  ... (last)")
        for ln in last_rows[-3:]:
            print(ln)


if __name__ == "__main__":
    main()
