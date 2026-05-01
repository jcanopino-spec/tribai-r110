"""Extrae la estructura del Formulario 110 desde la hoja '110 MUISCA'.

Patron observado:
- Encabezados de seccion en columna A (e.g., 'Datos informativos', 'Patrimonio',
  'Ingresos', 'Costos y deducciones', 'Renta', ...).
- Renglones con descripcion '<n>. <texto>' en cualquier columna y la formula en
  la celda inmediatamente debajo en la misma columna. La columna varia (B, D, F).

Salida: data/extracted/04_formulario_110.json
Estructura: lista ordenada de objetos:
    {numero, descripcion, seccion, formula_xlsm, fuente_celda}
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

SHEET = "110 MUISCA"
HEADER_RE = re.compile(r"^\s*(\d{1,3})\.\s+(.+?)\s*$", re.S)
SECTION_HEADERS = {
    "datos informativos",
    "datos del declarante",
    "patrimonio",
    "ingresos",
    "costos y deducciones",
    "esal (rte)",
    "esal (r.t.e.)",
    "renta",
    "ganancias ocasionales",
    "liquidacion privada",
    "liquidación privada",
    "signatarios",
    "compensaciones",
    "rentas exentas",
    "rentas gravables",
    "renta liquida gravable",
    "renta líquida gravable",
}


def is_section_header(value) -> str | None:
    if not isinstance(value, str):
        return None
    s = value.strip().lower()
    if s in SECTION_HEADERS:
        return value.strip()
    return None


def extract(xlsm_path: Path) -> list[dict]:
    wb = load_workbook(xlsm_path, data_only=False, keep_links=False)
    ws = wb[SHEET]

    results: list[dict] = []
    seen_numbers: set[int] = set()
    current_section: str | None = None

    # Build a 2D list for easy indexing (1-based via dict)
    grid: dict[tuple[int, int], object] = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
        for cell in row:
            v = cell.value
            if v is not None:
                grid[(cell.row, cell.column)] = v

    rows = sorted({r for (r, _) in grid})

    for r in rows:
        # Section detection (column A)
        a_val = grid.get((r, 1))
        sec = is_section_header(a_val)
        if sec:
            current_section = sec
            continue

        # Look in any column for "<num>. <text>"
        for c in range(2, ws.max_column + 1):
            v = grid.get((r, c))
            if not isinstance(v, str):
                continue
            m = HEADER_RE.match(v)
            if not m:
                continue
            num = int(m.group(1))
            desc = m.group(2).strip()
            # Heuristic: ignore plausible non-renglon numerics like "1. Año", but the form
            # actually has those — they are the field number 1.
            if num in seen_numbers:
                continue

            # Find formula in same column, next non-empty row
            formula_val = None
            formula_cell = None
            for r2 in (r2 for r2 in rows if r2 > r):
                fv = grid.get((r2, c))
                if fv is not None:
                    formula_val = fv
                    formula_cell = (r2, c)
                    break
                # Stop searching after we cross 5 rows of nothing in this col
                if r2 - r > 5:
                    break

            results.append({
                "numero": num,
                "descripcion": desc,
                "seccion": current_section,
                "formula_xlsm": str(formula_val) if formula_val is not None else None,
                "fuente_celda": f"{get_column_letter(c)}{formula_cell[0]}" if formula_cell else None,
                "fuente_celda_descripcion": f"{get_column_letter(c)}{r}",
            })
            seen_numbers.add(num)

    results.sort(key=lambda x: x["numero"])
    return results


def main():
    if len(sys.argv) != 2:
        print("usage: extract_form110.py <path>", file=sys.stderr)
        sys.exit(1)
    src = Path(sys.argv[1])
    out_dir = Path(__file__).resolve().parent.parent / "data" / "extracted"
    out_dir.mkdir(parents=True, exist_ok=True)
    rows = extract(src)
    (out_dir / "04_formulario_110.json").write_text(
        json.dumps(rows, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    print(f"Renglones detectados: {len(rows)}")
    if rows:
        print(f"Rango: {rows[0]['numero']}..{rows[-1]['numero']}")
    print()
    secciones: dict[str, int] = {}
    for r in rows:
        s = r.get("seccion") or "(sin seccion)"
        secciones[s] = secciones.get(s, 0) + 1
    for s, n in secciones.items():
        print(f"  {s:<35} {n} renglones")
    print()
    print("Primeros 5:")
    for r in rows[:5]:
        f = (r["formula_xlsm"] or "")[:60]
        print(f"  {r['numero']:>3}. {r['descripcion'][:55]:<55}  → {f}")
    print()
    print("Ultimos 5:")
    for r in rows[-5:]:
        f = (r["formula_xlsm"] or "")[:60]
        print(f"  {r['numero']:>3}. {r['descripcion'][:55]:<55}  → {f}")


if __name__ == "__main__":
    main()
