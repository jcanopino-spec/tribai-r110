"""Extrae catalogos limpios de la hoja 'Estructura detallada CIIU'.

Esa hoja agrupa varios catalogos:
- Col A-B-C → codigos CIIU (col C vacia) y codigos de regimen/tarifa (col C con tarifa)
- Col D-E   → codigo + nombre Direccion Seccional DIAN

Salidas:
    data/extracted/01_ciiu.json
    data/extracted/02_direcciones_seccionales.json
    data/extracted/03_regimenes_tarifas.json
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def is_int_str(s: str) -> bool:
    return s.lstrip("-").isdigit()


def extract(xlsm_path: Path) -> tuple[list[dict], list[dict], list[dict]]:
    wb = load_workbook(xlsm_path, data_only=True, keep_links=False)
    ws = wb["Estructura detallada CIIU"]

    ciiu: list[dict] = []
    dian: list[dict] = []
    regimenes: list[dict] = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        cells = [v for v in row]
        cells += [None] * (5 - len(cells))
        a, b, c, d, e = cells[:5]
        a_s = str(a).strip() if a is not None else ""
        b_s = str(b).strip() if b is not None else ""
        d_s = str(d).strip() if d is not None else ""
        e_s = str(e).strip() if e is not None else ""

        if is_int_str(a_s) and b_s and len(a_s) <= 4:
            if isinstance(c, (int, float)):
                # Codigo de regimen / tarifa
                regimenes.append({
                    "codigo": a_s.zfill(2),
                    "descripcion": b_s,
                    "tarifa": float(c),
                })
            else:
                ciiu.append({"codigo": a_s.zfill(4), "descripcion": b_s})

        if is_int_str(d_s) and e_s:
            dian.append({"codigo": d_s, "nombre": e_s})

    ciiu = list({(c["codigo"], c["descripcion"]): c for c in ciiu}.values())
    dian = list({(d["codigo"], d["nombre"]): d for d in dian}.values())
    regimenes = list({(r["codigo"], r["descripcion"]): r for r in regimenes}.values())
    ciiu.sort(key=lambda x: x["codigo"])
    dian.sort(key=lambda x: int(x["codigo"]))
    regimenes.sort(key=lambda x: int(x["codigo"]))

    return ciiu, dian, regimenes


def main():
    if len(sys.argv) != 2:
        print("usage: extract_catalogs.py <path>", file=sys.stderr)
        sys.exit(1)
    src = Path(sys.argv[1])
    out_dir = Path(__file__).resolve().parent.parent / "data" / "extracted"
    out_dir.mkdir(parents=True, exist_ok=True)

    ciiu, dian, regimenes = extract(src)
    (out_dir / "01_ciiu.json").write_text(
        json.dumps(ciiu, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    (out_dir / "02_direcciones_seccionales.json").write_text(
        json.dumps(dian, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    (out_dir / "03_regimenes_tarifas.json").write_text(
        json.dumps(regimenes, indent=2, ensure_ascii=False), encoding="utf-8"
    )

    print(f"CIIU:                       {len(ciiu)} actividades")
    print(f"Direcciones seccionales:    {len(dian)} administraciones DIAN")
    print(f"Regimenes/tarifas:          {len(regimenes)} categorias")
    print()
    print("CIIU primeros 3:")
    for c in ciiu[:3]:
        print(" ", c)
    print("Regimenes:")
    for r in regimenes:
        print(" ", r)


if __name__ == "__main__":
    main()
