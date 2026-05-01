"""Extrae el mapeo PUC -> Renglon 110 / Anexo / F-2516 desde 'Hoja Sumaria'.

Estructura observada (1-indexed):
- Col B: cuenta PUC (puede ser 1, 11, 1105, 110505, etc.)
- Col C: codigo del renglon en el formulario 110 (R-110)
- Col D: descripcion contable
- Col E: saldo contable (vacio en la plantilla)
- Col F-G: ajustes fiscales (debitos / creditos)
- Col H: saldo fiscal F-110
- Col I: anexo
- Col J: F-2516
- Col K: TTD (Tasa Tributaria Diferida)

Salida: data/extracted/05_hoja_sumaria.json
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def extract(xlsm_path: Path) -> list[dict]:
    wb = load_workbook(xlsm_path, data_only=True, keep_links=False)
    ws = wb["Hoja Sumaria"]

    rows: list[dict] = []
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        cells = list(row) + [None] * (11 - len(row))
        cuenta = cells[1]   # col B
        renglon = cells[2]  # col C
        desc = cells[3]     # col D
        anexo = cells[8]    # col I
        f2516 = cells[9]    # col J
        ttd = cells[10]     # col K

        # Skip rows with no PUC code
        if cuenta is None or (isinstance(cuenta, str) and not cuenta.strip()):
            continue

        # Normalize PUC to string of digits
        cuenta_str = str(cuenta).strip()
        if not cuenta_str.replace("-", "").isdigit():
            continue

        rows.append({
            "puc": cuenta_str,
            "renglon_110": int(renglon) if isinstance(renglon, (int, float)) else (renglon if renglon else None),
            "descripcion": str(desc).strip() if desc else "",
            "anexo": int(anexo) if isinstance(anexo, (int, float)) else (str(anexo).strip() if anexo else None),
            "f2516": int(f2516) if isinstance(f2516, (int, float)) else (str(f2516).strip() if f2516 else None),
            "ttd": str(ttd).strip() if ttd else None,
        })

    return rows


def main():
    src = Path(sys.argv[1])
    out_dir = Path(__file__).resolve().parent.parent / "data" / "extracted"
    out_dir.mkdir(parents=True, exist_ok=True)
    rows = extract(src)
    (out_dir / "05_hoja_sumaria.json").write_text(
        json.dumps(rows, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    print(f"Cuentas PUC: {len(rows)}")

    # Stats
    nivel_count: dict[int, int] = {}
    con_renglon = 0
    con_anexo = 0
    for r in rows:
        l = len(r["puc"])
        nivel_count[l] = nivel_count.get(l, 0) + 1
        if r["renglon_110"]:
            con_renglon += 1
        if r["anexo"]:
            con_anexo += 1

    print(f"Con renglón 110: {con_renglon}")
    print(f"Con anexo:       {con_anexo}")
    print("Niveles PUC (longitud del codigo → cantidad):")
    for l in sorted(nivel_count):
        print(f"  {l} dígitos: {nivel_count[l]}")
    print()
    print("Primeras 5:")
    for r in rows[:5]:
        print(f"  {r['puc']:<8} R{str(r['renglon_110'] or '-'):>3}  {r['descripcion'][:60]}")


if __name__ == "__main__":
    main()
