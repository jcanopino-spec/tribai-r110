#!/usr/bin/env python3
"""Sincroniza el mapeo PUC → renglón F110 con el del .xlsm guía v5.

El catálogo `puc_accounts` en BD tiene 3146 entradas pero algunas
están mal clasificadas vs la hoja "Detalle Fiscal" del archivo guía
oficial. Este script:

1. Lee el mapeo correcto del .xlsm (174 entradas oficiales)
2. Actualiza puc_accounts.renglon_110 para esos PUCs
3. Re-mapea balance_prueba_lineas.renglon_110 para todas las
   declaraciones afectadas
4. Recalcula form110_valores agregando por renglón

Idempotente: se puede correr múltiples veces.
"""

from pathlib import Path
import psycopg2
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSM = "/Users/jacp/Library/CloudStorage/OneDrive-Personal/CLAUDE/RENTA_110/R110/Tribai_R110_AG2025_v5_prueba_aries.xlsx"


def load_env() -> dict:
    env = {}
    for line in (ROOT / ".env.local").read_text().splitlines():
        if "=" in line and not line.startswith("#"):
            k, _, v = line.partition("=")
            env[k.strip()] = v.strip().strip("'\"")
    return env


def extract_xlsm_mapping() -> dict[str, int]:
    """Extrae mapping PUC → renglon del Detalle Fiscal del .xlsm."""
    wb = load_workbook(XLSM, data_only=False)
    ws = wb["Detalle Fiscal"]
    mapping: dict[str, int] = {}
    renglon_actual: int | None = None
    for r in range(12, ws.max_row + 1):
        rgl = ws.cell(r, 2).value
        puc = ws.cell(r, 3).value
        if rgl is not None:
            try:
                renglon_actual = int(rgl)
            except (ValueError, TypeError):
                renglon_actual = None
        if puc and renglon_actual is not None:
            puc_clean = str(puc).strip()
            # Saltar placeholders como (BIO1), (BIO2)
            if puc_clean.startswith("(") or not puc_clean.isdigit():
                continue
            # Saltar mapeos a R33-R35 (informativos · vienen de Anexo
            # Seguridad Social, NO del balance directamente). Esas cuentas
            # 5105-5195 deben ir a R63 según la fórmula SUMIF '51*' del
            # Detalle Fiscal del .xlsm guía.
            if renglon_actual in (33, 34, 35):
                continue
            mapping[puc_clean] = renglon_actual
    return mapping


def main():
    env = load_env()

    print("📖 Leyendo mapeo del .xlsm guía...")
    xlsm_map = extract_xlsm_mapping()
    print(f"   {len(xlsm_map)} PUCs oficiales (excluyendo R33-R35 informativos)")

    # Agregar reglas SUMIF por clase del .xlsm guía:
    #   R63 = SUMIF '51*'  · gastos administración (incluye 51xx personal)
    #   R64 = SUMIF '52*'  · gastos comercialización y ventas
    #   R65 = SUMIF '53*'  · gastos financieros
    #   R66 = SUMIF '54*'  · otros gastos
    # Estos prefijos de 2 dígitos son MENOS específicos que los del Detalle
    # Fiscal · solo aplican cuando NO hay mapeo más específico.
    # Como mi DISTINCT ON usa length DESC, los prefijos cortos se aplican
    # solo a cuentas que no matchearon ningún prefijo del .xlsm.
    xlsm_map.setdefault("51", 63)
    xlsm_map.setdefault("52", 64)
    xlsm_map.setdefault("53", 65)
    xlsm_map.setdefault("54", 66)
    print(f"   + reglas SUMIF clase 5: R63=51* · R64=52* · R65=53* · R66=54*")
    print(f"   total: {len(xlsm_map)} prefijos para resolver el balance")

    conn = psycopg2.connect(
        host="aws-1-sa-east-1.pooler.supabase.com",
        port=6543,
        database="postgres",
        user="postgres.wnbcdbfvriygtmodtytn",
        password=env["SUPABASE_DB_PASSWORD"],
        sslmode="require",
    )

    with conn.cursor() as cur:
        # 1. Estado antes
        cur.execute("SELECT count(*) FROM puc_accounts WHERE renglon_110 IS NOT NULL")
        before = cur.fetchone()[0]
        print(f"\n📊 puc_accounts con renglon_110 antes: {before}")

        # 2. Update por PUC oficial
        print("\n🔧 Actualizando puc_accounts según mapeo oficial...")
        cambios = 0
        nuevos = 0
        for puc, rgl in xlsm_map.items():
            cur.execute(
                """
                INSERT INTO puc_accounts (puc, renglon_110, ano_gravable)
                VALUES (%s, %s, 2025)
                ON CONFLICT (puc) DO UPDATE SET renglon_110 = EXCLUDED.renglon_110
                WHERE puc_accounts.renglon_110 IS DISTINCT FROM EXCLUDED.renglon_110
                """,
                (puc, rgl),
            )
            if cur.rowcount > 0:
                cambios += 1
        conn.commit()
        print(f"   {cambios} cuentas actualizadas")

        # 3. Re-mapear balance_prueba_lineas usando SOLO los 170 prefijos
        # oficiales del .xlsm (ignorando puc_accounts que tiene mapeos
        # contradictorios). Para cada cuenta del balance, encontramos el
        # prefijo MÁS LARGO del .xlsm que matchee · ese determina su renglón.
        # Replica la lógica SUMIF del .xlsm.
        print("\n🔧 Re-mapeando balance_prueba_lineas con prefijos oficiales...")
        # Construir VALUES con todos los prefijos
        values_sql = ",".join(
            f"('{puc}', {rgl})" for puc, rgl in xlsm_map.items()
        )
        cur.execute(f"""
            WITH prefijos(puc, rgl) AS (VALUES {values_sql}),
            matches AS (
                SELECT DISTINCT ON (l.id)
                    l.id, p.rgl
                FROM balance_prueba_lineas l
                JOIN prefijos p ON regexp_replace(l.cuenta, '[^0-9]', '', 'g') LIKE p.puc || '%%'
                ORDER BY l.id, length(p.puc) DESC
            )
            UPDATE balance_prueba_lineas l
            SET renglon_110 = m.rgl
            FROM matches m
            WHERE l.id = m.id
              AND l.renglon_110 IS DISTINCT FROM m.rgl
        """)
        afectadas = cur.rowcount
        conn.commit()
        print(f"   {afectadas} líneas re-mapeadas")

        # 4. Recalcular form110_valores agregando por renglón para cada
        # declaración con balance.
        print("\n🔧 Recalculando form110_valores...")
        cur.execute("""
            SELECT DISTINCT b.declaracion_id
            FROM balance_pruebas b
        """)
        decls = [r[0] for r in cur.fetchall()]
        print(f"   {len(decls)} declaraciones afectadas")

        for decl_id in decls:
            # Borrar valores actuales
            cur.execute("DELETE FROM form110_valores WHERE declaracion_id = %s", (decl_id,))

            # Re-insertar agregando por renglón. Aplicar:
            #   1. Filtro anti-duplicación: solo cuentas hoja (sin hijas)
            #   2. Saldo fiscal: saldo + ajuste_debito - ajuste_credito
            #      Esto es consistente con loadF2516Aggregates (que aplica
            #      la misma fórmula). Si el balance no tiene ajustes capturados,
            #      saldo_fiscal = saldo (compat con balances simples).
            #   3. Normalización de signo: pasivos/ingresos vienen como
            #      saldo crédito (negativo), los pasamos a positivo.
            cur.execute("""
                WITH lineas AS (
                    SELECT
                        regexp_replace(l.cuenta, '[^0-9]', '', 'g') as cuenta_num,
                        l.renglon_110,
                        l.saldo + COALESCE(l.ajuste_debito, 0) - COALESCE(l.ajuste_credito, 0) as saldo_fiscal,
                        b.id as balance_id
                    FROM balance_prueba_lineas l
                    JOIN balance_pruebas b ON b.id = l.balance_id
                    WHERE b.declaracion_id = %s
                      AND l.renglon_110 IS NOT NULL
                ),
                hojas AS (
                    -- Filtrar cuentas resumen: solo cuentas SIN hijas en el balance
                    SELECT l.*
                    FROM lineas l
                    WHERE NOT EXISTS (
                        SELECT 1 FROM lineas l2
                        WHERE l2.balance_id = l.balance_id
                          AND length(l2.cuenta_num) > length(l.cuenta_num)
                          AND l2.cuenta_num LIKE l.cuenta_num || '%%'
                    )
                ),
                agregado AS (
                    SELECT
                        renglon_110,
                        sum(saldo_fiscal) as suma_natural,
                        -- Pasivos (R45) e ingresos (R47-R57) tienen naturaleza crédito
                        -- → su suma natural es negativa. Convertimos a positivo.
                        CASE
                            WHEN renglon_110 = 45 THEN abs(sum(saldo_fiscal))
                            WHEN renglon_110 BETWEEN 47 AND 57 THEN abs(sum(saldo_fiscal))
                            WHEN renglon_110 = 59 THEN abs(sum(saldo_fiscal))
                            WHEN renglon_110 = 60 THEN abs(sum(saldo_fiscal))
                            ELSE sum(saldo_fiscal)
                        END as valor_normalizado
                    FROM hojas
                    GROUP BY renglon_110
                )
                INSERT INTO form110_valores (declaracion_id, numero, valor)
                SELECT %s, renglon_110, valor_normalizado
                FROM agregado
                WHERE valor_normalizado != 0
            """, (decl_id, decl_id))
            inserted = cur.rowcount
            print(f"   decl {decl_id[:8]}... · {inserted} renglones recalculados")

        conn.commit()

        # 5. Estado final
        cur.execute("SELECT count(*) FROM puc_accounts WHERE renglon_110 IS NOT NULL")
        after = cur.fetchone()[0]
        print(f"\n📊 puc_accounts con renglon_110 después: {after}")

    conn.close()
    print("\n✓ Fix completo.")


if __name__ == "__main__":
    main()
